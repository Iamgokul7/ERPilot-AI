import io
import csv
from flask import Blueprint, render_template, request, Response, send_file, session, redirect, url_for, flash
from utils.auth_utils import role_required
from utils.db_utils import require_tenant_id
from database import get_connection

# ReportLab imports
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors

# OpenPyXL imports
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

reports_bp = Blueprint("reports", __name__, url_prefix="/reports")

@reports_bp.route("")
@role_required(['admin', 'manager', 'employee'])
def reports_dashboard():
    tenant_id = require_tenant_id()
    conn = get_connection()
    cursor = conn.cursor()
    
    # Get stats for reports page context
    cursor.execute("SELECT COUNT(*) FROM employees WHERE tenant_id = ?", (tenant_id,))
    emp_count = cursor.fetchone()[0]
    cursor.execute("SELECT COUNT(*) FROM products WHERE tenant_id = ?", (tenant_id,))
    prod_count = cursor.fetchone()[0]
    cursor.execute("SELECT COUNT(*) FROM suppliers WHERE tenant_id = ?", (tenant_id,))
    supp_count = cursor.fetchone()[0]
    cursor.execute("SELECT COUNT(*) FROM sales_orders WHERE tenant_id = ?", (tenant_id,))
    sales_count = cursor.fetchone()[0]
    
    # Get company name
    company_name = "Acme Enterprise Solutions"
    cursor.execute("SELECT company_name FROM tenants WHERE id = ?", (tenant_id,))
    tenant = cursor.fetchone()
    if tenant:
        company_name = tenant["company_name"]
        
    conn.close()
    
    return render_template(
        "reports/reports.html",
        emp_count=emp_count,
        prod_count=prod_count,
        supp_count=supp_count,
        sales_count=sales_count,
        company_name=company_name
    )

@reports_bp.route("/export")
@role_required(['admin', 'manager', 'employee'])
def export_report():
    tenant_id = require_tenant_id()
    report_type = request.args.get("type", "inventory").strip()
    export_format = request.args.get("format", "csv").strip()
    
    conn = get_connection()
    cursor = conn.cursor()
    
    data = []
    headers = []
    title = ""
    
    # Fetch Data based on Type
    if report_type == "employees":
        title = "Employee Report"
        headers = ["Employee ID", "Name", "Email", "Phone", "Department", "Designation", "Join Date", "Status"]
        cursor.execute("""
            SELECT employee_id, name, email, phone, department, designation, join_date, status
            FROM employees WHERE tenant_id = ? ORDER BY id DESC
        """, (tenant_id,))
        data = [list(row) for row in cursor.fetchall()]
        
    elif report_type == "inventory":
        title = "Inventory Report"
        headers = ["Product Name", "SKU", "Category", "Quantity", "Unit Price", "Reorder Level"]
        cursor.execute("""
            SELECT product_name, sku, category, quantity, unit_price, reorder_level
            FROM products WHERE tenant_id = ? ORDER BY id DESC
        """, (tenant_id,))
        data = [list(row) for row in cursor.fetchall()]
        
    elif report_type == "suppliers":
        title = "Supplier Report"
        headers = ["Supplier Name", "Contact Person", "Email", "Phone", "Address"]
        cursor.execute("""
            SELECT supplier_name, contact_person, email, phone, address
            FROM suppliers WHERE tenant_id = ? ORDER BY id DESC
        """, (tenant_id,))
        data = [list(row) for row in cursor.fetchall()]
        
    elif report_type == "sales":
        title = "Sales Orders Report"
        headers = ["Order Number", "Customer Name", "Product", "Quantity", "Total Amount", "Order Date", "Status"]
        cursor.execute("""
            SELECT o.order_number, o.customer_name, p.product_name, o.quantity, o.total_amount, o.order_date, o.status
            FROM sales_orders o
            LEFT JOIN products p ON o.product_id = p.id
            WHERE o.tenant_id = ? ORDER BY o.id DESC
        """, (tenant_id,))
        data = [list(row) for row in cursor.fetchall()]
        
    else:
        conn.close()
        flash("Invalid report type requested.", "danger")
        return redirect(url_for("reports.reports_dashboard"))
        
    conn.close()
    
    # Handle Exports
    if export_format == "csv":
        return export_csv(title, headers, data)
    elif export_format == "excel":
        return export_excel(title, headers, data)
    elif export_format == "pdf":
        return export_pdf(title, headers, data)
    else:
        flash("Invalid export format requested.", "danger")
        return redirect(url_for("reports.reports_dashboard"))

def export_csv(title, headers, data):
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([title.upper()])
    writer.writerow([])
    writer.writerow(headers)
    for row in data:
        writer.writerow(row)
    
    response = Response(output.getvalue(), mimetype="text/csv")
    filename = f"{title.lower().replace(' ', '_')}.csv"
    response.headers["Content-Disposition"] = f"attachment; filename={filename}"
    return response

def export_excel(title, headers, data):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title[:30]
    
    # Styles
    title_font = Font(name="Arial", size=16, bold=True, color="000000")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="111111", end_color="111111", fill_type="solid")
    data_font = Font(name="Arial", size=10)
    
    ws.append([title])
    ws.cell(1, 1).font = title_font
    ws.append([]) # empty row
    
    ws.append(headers)
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(3, col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        
    for row in data:
        ws.append(row)
        
    # Autofit columns
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    
    filename = f"{title.lower().replace(' ', '_')}.xlsx"
    return send_file(
        out,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename
    )

def export_pdf(title, headers, data):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        rightMargin=30,
        leftMargin=30,
        topMargin=30,
        bottomMargin=30
    )
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Custom Styles (Monochrome)
    title_style = ParagraphStyle(
        'ReportTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=18,
        textColor=colors.HexColor('#111111'),
        spaceAfter=15
    )
    
    elements.append(Paragraph(title, title_style))
    elements.append(Spacer(1, 10))
    
    # Process text cells in table data to fit columns neatly
    table_data = [headers]
    for row in data:
        row_data = []
        for cell in row:
            # Format numbers cleanly
            if isinstance(cell, float):
                row_data.append(f"${cell:,.2f}" if "price" in title.lower() or "amount" in title.lower() or "revenue" in title.lower() else f"{cell}")
            else:
                row_data.append(str(cell if cell is not None else ''))
        table_data.append(row_data)
        
    t = Table(table_data)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#111111')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0,0), (-1,0), 8),
        ('BACKGROUND', (0,1), (-1,-1), colors.HexColor('#FAFAFA')),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E5E5E5')),
        ('FONTNAME', (0,1), (-1,-1), 'Helvetica'),
        ('FONTSIZE', (0,0), (-1,-1), 8),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#F9F9F9')]),
        ('TOPPADDING', (0,0), (-1,-1), 6),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
    ]))
    
    elements.append(t)
    doc.build(elements)
    
    buffer.seek(0)
    filename = f"{title.lower().replace(' ', '_')}.pdf"
    return send_file(
        buffer,
        mimetype="application/pdf",
        as_attachment=True,
        download_name=filename
    )
